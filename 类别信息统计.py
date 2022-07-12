# -*- coding:utf-8 -*-
from tkinter.messagebox import showinfo

import openpyxl
import datetime
import tkinter as tk
from tkinter import filedialog
window = tk.Tk()
window.title('歌词信息提取')
window.geometry('500x300')
window.config(background = '#d6d6d6')
import os


lie1 = 20
lie2 = 190

tk.Label(window, text="请输入歌词在表格内的列数：",bg = "#d6d6d6").place(x=lie1, y=30)
chaifen_cout = tk.StringVar()  # 文件输入路径变量

tk.Label(window, text="请输入工作表名称：",bg = "#d6d6d6").place(x=lie1, y=70)
chaifen_cout2 = tk.StringVar()  # 文件输入路径变量

tk.Label(window, text="请选择类别表数据文件：",bg = "#d6d6d6").place(x=lie1, y=110)
var_name = tk.StringVar()  # 文件输入路径变量

tk.Label(window, text="请选择文件解析后存储位置：",bg = "#d6d6d6").place(x=lie1, y=150)
var_name2 = tk.StringVar()  # 文件夹输入路径变量

tk.Label(window, text="请选择文件解析后存储位置：",bg = "#d6d6d6").place(x=lie1, y=190)
var_name3 = tk.StringVar()  # 文件夹输入路径变量

entry_name = tk.Entry(window, textvariable=chaifen_cout, width=20)
entry_name.place(x=lie2, y=30)

entry_name5 = tk.Entry(window, textvariable=chaifen_cout2, width=20)
entry_name5.place(x=lie2, y=70)

entry_name = tk.Entry(window, textvariable=var_name, width=25)
entry_name.place(x=lie2, y=110)

entry_name2 = tk.Entry(window, textvariable=var_name2, width=25)
entry_name2.place(x=lie2, y=150)

entry_name3 = tk.Entry(window, textvariable=var_name3, width=25)
entry_name3.place(x=lie2, y=190)

def selectPath_file():#文件选择
    path_ = filedialog.askopenfilename(filetypes=[("数据表", [".xlsx"])])
    print(path_)
    var_name.set(path_)

# 输入文件夹路径
def selectPath_dir():
    path_t = filedialog.askdirectory()
    print(path_t)
    var_name2.set(path_t)

def selectPath_dir2():
    path_t2 = filedialog.askdirectory()
    print(path_t2)
    var_name3.set(path_t2)

def mains():
    start = datetime.datetime.now()
    path = var_name.get()
    path3 = var_name2.get()
    for i, k, l in os.walk(path3):
        for l2 in l:
            print(l2)
            path1 = path3 + '/' + l2
    path2 = var_name2.get()

    numa = int(chaifen_cout.get())
    she = chaifen_cout2.get()
    wb = openpyxl.load_workbook(path)
    sheet = wb[she]
    row_num = sheet.max_row
    column_num = sheet.max_column
    fenlei_diss  = []
    fenlei_diss2 = []


    for i in range(numa+1,column_num+1):
        all_data_list = []
        not_all_list = []
        all_data_diss = {}
        all_data_diss2 = {}
        biaotou = sheet.cell(1, i).value
        for j in  range(2,row_num+1):
            xinxi = sheet.cell(j,i).value
            if xinxi !='无' or xinxi == None:
                all_data_list.append(xinxi)
                if xinxi not in not_all_list:
                    not_all_list.append(xinxi)
                else:
                    continue
            else:
                continue
        all_data_diss['biaotou'] = biaotou
        all_data_diss['quanbushuliang'] = len(all_data_list)
        all_data_diss['xinxi'] =  all_data_list
        fenlei_diss.append(all_data_diss)

        all_data_diss2['biaotou'] = biaotou
        all_data_diss2['quchonghoudeshuliang'] = len(not_all_list)
        all_data_diss2['xinxi'] = not_all_list

        fenlei_diss2.append(all_data_diss2)
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.create_sheet('数据未去重')
    sheet3 = wb2.create_sheet('数据已去重')
    for nums in range(len(fenlei_diss)):#类别
        print(fenlei_diss[nums])
        print('---------------------------------------------------')
        sheet2.cell(1,nums+1).value=fenlei_diss[nums]['biaotou']
        sheet2.cell(2,nums+1).value=fenlei_diss[nums]['quanbushuliang']
        for nums_2 in range(len(fenlei_diss[nums]['xinxi'])):
            sheet2.cell(nums_2+3, nums + 1).value = fenlei_diss[nums]['xinxi'][nums_2]

    for nums2 in range(len(fenlei_diss2)):#类别
        print(fenlei_diss2[nums2])
        print('---------------------------------------------------')
        sheet3.cell(1,nums2+1).value=fenlei_diss2[nums2]['biaotou']
        sheet3.cell(2, nums2 + 1).value = fenlei_diss2[nums2]['quchonghoudeshuliang']
        for nums_2 in range(len(fenlei_diss2[nums2]['xinxi'])):
            sheet3.cell(nums_2+3, nums2 + 1).value = fenlei_diss2[nums2]['xinxi'][nums_2]
    wb2.save(path2 + '/类别信息筛选表.xlsx')
    end = datetime.datetime.now()
    showinfo('提示', '运行完毕！' + '\n' + '程序运行时间: ' + str(((end - start) / 60).seconds) + '分钟')
    window.quit()

# def mains2():
#     path3 = 'C:/Users/zhihaimao/Desktop/123/222'
#     for i, k, l in os.walk(path3):
#         for l2 in l:
#             print(l2)
#             path1 = path3 + '/' + l2
#             mains(path1)
tk.Button(window, text="文件选择", command=selectPath_file,bg = '#d1d1d1').place(x=400, y=105)
tk.Button(window, text="请选择文件夹", command=selectPath_dir,bg = '#d1d1d1').place(x=400, y=150)
tk.Button(window, text="请选择文件夹", command=selectPath_dir2,bg = '#d1d1d1').place(x=400, y=190)
tk.Button(window, text="运行", command=mains,width = 12,height = 1,bg = '#d1d1d1').place(x=210, y=230)
window.mainloop()  # 显示窗口