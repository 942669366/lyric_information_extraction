import export as export

from tkinter.messagebox import *
import  tkinter.messagebox
import openpyxl
import re
import datetime
from collections import Counter
import tkinter as tk
from tkinter import filedialog
import os
import _thread
import threading

window = tk.Tk()
window.title('歌词信息提取')
window.geometry('500x250')
window.config(background = '#d6d6d6')


lie1 = 20
lie2 = 190

tk.Label(window, text="请输入数据在表格内的列数：",bg = "#d6d6d6").place(x=lie1, y=30)
chaifen_cout = tk.StringVar()  # 文件输入路径变量

tk.Label(window, text="请选择类别表数据文件：",bg = "#d6d6d6").place(x=lie1, y=70)
var_name = tk.StringVar()  # 文件输入路径变量

tk.Label(window, text="选择拆分文件后存储位置：",bg = "#d6d6d6").place(x=lie1, y=110)
var_name2 = tk.StringVar()  # 文件夹输入路径变量

# tk.Label(window, text="请输入类别表数据所在的列数：",bg = "#d6d6d6").place(x=lie1, y=110)
# chaifen_cout2 = tk.StringVar()  # 文件输入路径变量

# tk.Label(window, text="请选择类别表数据文件：",bg = "#d6d6d6").place(x=lie1, y=110)
# var_name2 = tk.StringVar()  # 文件输入路径变量

# tk.Label(window, text="注意：1、所选文件内必须包含“Sheet1”和“Sheet2”两个工作表。",bg = "#d6d6d6",fg = '#FF0000').place(x=lie1, y=130)
# tk.Label(window, text="2、歌词必须在“Sheet1”工作表内。",bg = "#d6d6d6",fg = '#FF0000').place(x=lie1+36, y=150)
# tk.Label(window, text="3、执行程序运行前，需关闭执行的文件的表。",bg = "#d6d6d6",fg = '#FF0000').place(x=lie1+36, y=170)
entry_name = tk.Entry(window, textvariable=chaifen_cout, width=20)
entry_name.place(x=lie2, y=30)

entry_name = tk.Entry(window, textvariable=var_name, width=25)
entry_name.place(x=lie2, y=70)


# entry_name2 = tk.Entry(window, textvariable=chaifen_cout2, width=20)
# entry_name2.place(x=lie2, y=110)

entry_name2 = tk.Entry(window, textvariable=var_name2, width=25)
entry_name2.place(x=lie2, y=110)
# 输入文件路径
def selectPath_file():#文件选择
    path_ = filedialog.askopenfilename(filetypes=[("数据表", [".xlsx"])])
    print(path_)
    var_name.set(path_)

# def selectPath_file2():#文件选择
#     path_ = filedialog.askopenfilename(filetypes=[("数据表", [".xlsx"])])
#     print(path_)
#     var_name2.set(path_)

# 输入文件夹路径
def selectPath_dir():
    path_t = filedialog.askdirectory()
    print(path_t)
    var_name2.set(path_t)

def zuoci(geci,h):
    # print(geci)
    zuoci_list = ['作词 ','作词 :','作词 ：','作 词：','作 词:','词曲 ','词/ ','作词：','作词:','词/曲','词／曲','词lyricist','词Lyrics','作词Lyrics:','词曲/编曲：','词曲/编曲:','词曲/演唱：','词曲/演唱:','词曲唱','作词作曲演唱','词：','词:','词 ','Lyrics by ','Lyrics by：','Written by ','Written by：','詞 ','詞：','詞:','作詞 ','作詞：','作詞:','작사：','작사:','작사가：','작사가:','作詞する．歌詞を作る：','作詞する．歌詞を作る:','Text von：','Text von:','авторов：','авторов:','автор текста：','автор текста:']
    zuoci2 = ''
    for i in zuoci_list:
        if i in geci:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                zuoci2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
                # print(str(h) + geci)
            else:
                zuoci2 = '------------------------------------------------'
            break
        else:
            zuoci2 = '无'
    return zuoci2

def zuoqu(zuoqu1,h):
    zuoqu_list = ['词曲 ','词/曲','词／曲','作曲/ ','作曲 ','作曲:','作曲：','作 曲：','曲/后期','作曲Composer:','词曲/编曲：','词曲/编曲:','词曲/演唱：','词曲/演唱:', '词曲/演唱',',词曲、演唱','作词作曲演唱','词曲唱','作曲/编曲/混音', '作曲\编曲：','作曲\编曲:','曲：','曲:', '曲 ','Composed by ', 'Composed by：', '작곡：', '작곡:', '작곡가：', '작곡가:', 'さっきょく：','さっきょく:', 'komponist：', 'komponist:', 'композитор：', 'композитор:']
    zuoqu2 = ''
    for i in zuoqu_list:
        if i in zuoqu1:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), zuoqu1)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                zuoqu2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
            else:
                zuoqu2 = '------------------------------------------------'
            break
        else:
            zuoqu2 = '无'
    return zuoqu2

def bianqu(bianqu1,h):
    bianqu_list = ['编曲：', '编曲:', '编曲人:', '编曲人：', '编曲 ', '编曲 录音 后期','作曲\编曲：','作曲\编曲:','词曲/编曲：','词曲/编曲:','编曲/混音：', '编曲/混音:','编曲Music Arranger','编曲Arranger,','作曲/编曲/混音', '編曲：', '編曲:', 'Arranged by：', 'Arranged by:', '편곡자：', '편곡자:','편곡：', '편곡:', 'アレンジ：', 'アレンジ:', 'Arrangement:','Arrangement：', 'Аранжировка：', 'Аранжировка:']
    bianqu2 = ''
    for i in bianqu_list:
        if i in bianqu1:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), bianqu1)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                bianqu2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
                # print(str(h) + geci)
            else:
                bianqu2 = '------------------------------------------------'
            break
        else:
            bianqu2 = '无'
    return bianqu2

def zhizuoren(zhizuoren1,h):
    zhizuoren_list = ['制作人：', '制作人:','制 作 人：', '制作:', '制作：', '制作 ', '製作人:', '製作人：','制作人Music producer','制作人Producer', 'Produced by:', 'Produced by：', 'Produced：', 'Produced:','제작자:', '제작자：', '프로듀서:', '프로듀서：', 'プロデューサー：', 'プロデューサー:', 'Produzent：','Produzent:','изготовитель：','изготовитель:']
    zhizuoren2 = ''
    for i in zhizuoren_list:
        if i in zhizuoren1:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), zhizuoren1)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                zhizuoren2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
                # print(str(h) + geci)
            else:
                zhizuoren2 = '------------------------------------------------'
            break
        else:
            zhizuoren2 = '无'
    return zhizuoren2

def hunying(hunying1,h):
    hunying_list = ['混音师：','混音师:','录混 ','混音：','混音:','混 音：','混音/后期','混音\母带','编曲/混音：','录音混音：','录音混音:', '编曲/混音:','混 音 / 母 带','录音/混音工程师','混音/母带工程', '录音/混音','混音工程师 Mixing Engineer','混音师Mixing Engineer','人声录音/混音/母带/协调','作曲/编曲/混音', '录音/混音/母带：','录音/混音/母带:','录音Recording/混音Mixing','錄音：', '錄音:', '錄音師:', '錄音師：', 'Recording Engineer：','Recording Engineer:','녹음 기사:', '녹음 기사：', '녹음:', '녹음：', 'モノラル：', 'モノラル:', 'ろくおん：', 'ろくおん:','テープ付き：', 'テープ付き:','Tontechniker：','Tontechniker:','звукооператор：','звукооператор:']
    hunying2 = ''
    for i in hunying_list:
        if i in hunying1:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), hunying1)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                hunying2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
            else:
                hunying2 = '------------------------------------------------'
            break
        else:
            hunying2 = '无'
    return hunying2

def luying(luying1):
    luying_list = ['录音师：','录音师:','录音：', '录音:','录 音', '录混 ', '录音混音：','录音混音:', '录音/混音：', '录音/混音:', '混音工程师：', '混音工程师:','录音 Recording Engineer','编曲 录音 后期','录音/混音工程师','录音/混音/母带：','录音师Recording Engineer','录音师Recording engineer','录音Recording/混音Mixing', 'Mixing Engineer:','Mixing Engineer：', '사운드 믹서:', '사운드 믹서：', '리 믹스:', '리 믹스：', 'ミックスし師：', 'ミックスし師:', 'sound mixer:','sound mixer：','Ремиксы：', 'Ремиксы:']
    luying2 = ''
    for i in luying_list:
        if i in luying1:
            searchObj = re.findall(r'{}(.*?)%0A'.format(i), luying1)
            # return str(h) +str(searchObj)
            if len(searchObj) > 0:
                luying2 = searchObj[0].strip(' ').strip(':').strip('：').strip('%0D').strip('╱')
            else:
                luying2 = '------------------------------------------------'
            break
        else:
            luying2 = '无'
    return luying2

def bumber_panduan(n): #返回Ture表明字符串中包含数字
    bumber_panduan_aa = ''
    for i in n:
        if i.isdigit() == True:
            bumber_panduan_aa = True
            break
        else:
            bumber_panduan_aa = False
    return bumber_panduan_aa


def guileis(sheet,row_max):

    guilei_list = []
    for i in range(row_max):
        guilei = sheet.cell(i+2,4).value
        if guilei == None or '无意义' in guilei:
            continue
        else:
            if guilei not in guilei_list:

                guilei_list.append(guilei)
            else:
                continue

    wuxiegang = []

    for guilei2 in guilei_list:
        if '/' not in guilei2:
            wuxiegang.append(guilei2)
        else:
            guilei2_1 = guilei2.split('/')
            for guilei2_1_1 in guilei2_1:
                if guilei2_1_1 not in wuxiegang:
                    wuxiegang.append(guilei2_1_1)
                else:
                    continue
    wuxiegang2 = []
    for guilei4 in wuxiegang:
        if '／' in guilei4:
            guilei4_1 = guilei4.split('／')
            for guilei4_2 in guilei4_1:
                if guilei4_2 not in wuxiegang2:
                    wuxiegang2.append(guilei4_2)
                else:
                    continue
        else:
            if guilei4 not in wuxiegang2:
                wuxiegang2.append(guilei4)
            else:
                continue
    out_ls = ['编曲','曲作者','词曲作者','作曲','词曲','原曲作者','原词作者',' 混音师','作曲者','词作者','词','混音师','混音','录音师','录音','制作人','制作','','男声']#需要排除的关键类别
    for fi in out_ls:
        if fi in wuxiegang2:
            wuxiegang2.remove(fi)
        else:
            continue
    return wuxiegang2

def guileis2(sheet,row_max):
    wuxiegang2 = guileis(sheet,row_max)
    leibie1_dis = {}
    for leibie1 in wuxiegang2:
        leibie1_list = []
        for i1 in range(2,row_max+1):
            tiqu_data = sheet.cell(i1,2).value
            leibie1_1 = sheet.cell(i1,4).value
            if leibie1_1 == '无意义' or leibie1_1 == '无意义‘' or leibie1_1 == None:
                continue
            else:
                if leibie1 in leibie1_1:
                    leibie1_list.append(tiqu_data)
                else:
                    continue
        leibie1_dis[leibie1] = leibie1_list
    return leibie1_dis ,wuxiegang2

def all_leibie(id,all_dis,geci):
    fenlei_all_dis ={}
    for leibie_key,leibie_value in all_dis.items():
        zuoci2 = "无"
        for i in leibie_value:
            if i+':' in geci:
                searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
                if len(searchObj) > 0:
                    # print(searchObj)
                    zuoci2 = str(searchObj[0]).strip(' ').strip(':').strip('：').strip('%0D').strip('╱').strip('】').strip('。')
                    # print(str(leibie_key)+'----------11111----------------'+str(zuoci2))
                    break
                else:
                    zuoci2 = '----------------------'
                    break
            elif i+ '：' in geci:
                searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
                if len(searchObj) > 0:
                    # print(searchObj)
                    zuoci2 = str(searchObj[0]).strip(' ').strip(':').strip('：').strip('%0D').strip('╱').strip('】').strip('。')
                    # print(str(leibie_key)+'----------222222----------------'+str(zuoci2))
                    break
                else:
                    zuoci2 = '----------------------'
                    break
            elif i+ '/' in geci:
                searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
                if len(searchObj) > 0:
                    # print(searchObj)
                    zuoci2 = str(searchObj[0]).strip(' ').strip(':').strip('：').strip('%0D').strip('╱').strip('】').strip('。')
                    # print(str(leibie_key)+'-----------33333---------------'+str(zuoci2))
                    break
                else:
                    zuoci2 = '----------------------'
                    break
            elif i+ '／'in geci:
                searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
                if len(searchObj) > 0:
                    # print(searchObj)
                    zuoci2 = str(searchObj[0]).strip(' ').strip(':').strip('：').strip('%0D').strip('╱').strip('】').strip('。')
                    # print(str(leibie_key)+'-----------44444---------------'+str(zuoci2))
                    break
                else:
                    zuoci2 = '----------------------'
                    print(zuoci2)
                    break
            elif i+' ' in geci:
                searchObj = re.findall(r'{}(.*?)%0A'.format(i), geci)
                if len(searchObj) > 0:
                    # print(searchObj)
                    zuoci2 = str(searchObj[0]).strip(' ').strip(':').strip('：').strip('%0D').strip('╱').strip('】').strip('。')
                    break
                else:
                    zuoci2 = '----------------------'
                    break

            else:

                continue
        fenlei_all_dis[leibie_key] = zuoci2
    return fenlei_all_dis

def mains(path1,path3):
    wb = openpyxl.load_workbook(path1)
    sheet = wb['Sheet']
    row_num = sheet.max_row
    wb2 = openpyxl.load_workbook(path3)
    sheet2 = wb2['Sheet1']
    max_dis2 = sheet2.max_row
    all_dis,wuxiegang2 = guileis2(sheet2,max_dis2)

    line = int(chaifen_cout.get())
    zuoci_list = []
    zuoqu_list = []
    bianqu_list = []
    zhizuoren_list = []
    hunying_list = []
    luying_list = []
    null_cout = 0
    sheet.cell(1, 3).value = '作词'
    sheet.cell(1, 4).value = '作曲'
    sheet.cell(1, 5).value = '编曲'
    sheet.cell(1, 6).value = '制作人'
    sheet.cell(1, 7).value = '混音'
    sheet.cell(1, 8).value = '录音'
    for o in range(len(wuxiegang2)):
        sheet.cell(1, o + 9).value = wuxiegang2[o]
    for i in range(2,row_num+1):
        try:
            geci = str(sheet.cell(i,line).value)
            id = str(sheet2.cell(i, 1).value)
            if geci != None:
                zuoci_jieguo = zuoci(geci, i)
                zuoqu_jieguo = zuoqu(geci,i)
                bianqu_jieguo = bianqu(geci,i)
                zhizuoren_jieguo = zhizuoren(geci,i)
                hunying_jieguo = hunying(geci,i)
                luying_jieguo = luying(geci)
                fenlei_all_dis = all_leibie(id, all_dis, geci)
                for wuxiegang_one in range(len(wuxiegang2)):
                    sheet.cell(i, wuxiegang_one + 9).value = str(fenlei_all_dis[wuxiegang2[wuxiegang_one]])
                    print(str(wuxiegang2[wuxiegang_one]) + '+++++++++++++++++++++++++++++++++' + str(fenlei_all_dis[wuxiegang2[wuxiegang_one]]))
                print(len(fenlei_all_dis))
                # print(str(id) + '--------------------------------------------------------------------------------------------------------------------------------')

                sheet.cell(i, line+1).value = zuoci_jieguo
                sheet.cell(i, line+2).value = zuoqu_jieguo
                sheet.cell(i, line+3).value = bianqu_jieguo
                sheet.cell(i, line+4).value = zhizuoren_jieguo
                sheet.cell(i, line+5).value = hunying_jieguo
                sheet.cell(i, line+6).value = luying_jieguo

                zuoci_list.append(zuoci_jieguo)
                zuoqu_list.append(zuoqu_jieguo)
                bianqu_list.append(bianqu_jieguo)
                zhizuoren_list.append(zhizuoren_jieguo)
                hunying_list.append(hunying_jieguo)
                luying_list.append(luying_jieguo)
                print(str(i) + '--------------------' + str(zuoci_jieguo) + '--------------------' + str(zuoqu_jieguo) + '--------------------' + str(bianqu_jieguo) + '--------------------' + str(zhizuoren_jieguo) + '--------------------' + str(hunying_jieguo) + '--------------------' + str(luying_jieguo))
            else:
                null_cout += 1
                sheet.cell(i, line + 1).value = None
                sheet.cell(i, line + 2).value = None
                sheet.cell(i, line + 3).value = None
                sheet.cell(i, line + 4).value = None
                sheet.cell(i, line + 5).value = None
                sheet.cell(i, line + 6).value = None
        except Exception as err:
            print(err)
            continue
    wb.save(path1)


def mains2():
    start = datetime.datetime.now()
    path3 = var_name.get()
    path = var_name2.get()

    for i, k, l in os.walk(path):
        for l2 in l:
            path1 = path + '/' + l2
            mains(path1,path3)
    end = datetime.datetime.now()
    showinfo('提示', '运行完毕！' + '\n' + '程序运行时间: ' + str(((end - start) / 60).seconds) + '分钟')
    window.quit()

tk.Button(window, text="文件选择", command=selectPath_file,bg = '#d1d1d1').place(x=400, y=65)
tk.Button(window, text="请选择文件夹", command=selectPath_dir,bg = '#d1d1d1').place(x=400, y=105)
tk.Button(window, text="运行", command=mains2,width = 12,height = 1,bg = '#d1d1d1').place(x=210, y=200)
window.mainloop()  # 显示窗口
